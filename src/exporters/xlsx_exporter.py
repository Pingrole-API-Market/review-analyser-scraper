import io
import logging

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_SUMMARY_HEADERS = [
    "Platform", "Business", "Overall Rating", "Total Reviews",
    "Reviews (12 mo)", "5★", "4★", "3★", "2★", "1★",
]
_REVIEW_HEADERS = [
    "Platform", "Business", "Author", "Country", "Author Reviews",
    "Rating", "Title", "Body", "Date", "Verified", "Unprompted",
]
_REVIEW_KEYS = [
    "platform", "business_name", "author", "author_country", "author_total_reviews",
    "rating", "title", "body", "date", "verified", "unprompted",
]


def export_xlsx(results: list[dict]) -> bytes:
    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_header_row(ws_sum, _SUMMARY_HEADERS)
    for r in results:
        bd = r.get("rating_breakdown") or {}
        ws_sum.append([
            r.get("platform"),
            r.get("business_name"),
            r.get("overall_rating"),
            r.get("total_reviews"),
            r.get("reviews_last_12_months"),
            bd.get("5_star"), bd.get("4_star"), bd.get("3_star"),
            bd.get("2_star"), bd.get("1_star"),
        ])
    _autowidth(ws_sum)

    # ── One Reviews sheet per platform ───────────────────────────────
    for result in results:
        platform = (result.get("platform") or "reviews")[:31]  # Excel sheet name limit
        ws = wb.create_sheet(platform.capitalize())
        _write_header_row(ws, _REVIEW_HEADERS)

        base = {
            "platform":      result.get("platform", ""),
            "business_name": result.get("business_name", ""),
        }
        for review in result.get("reviews", []):
            row_data = {**base, **review}
            ws.append([row_data.get(k, "") for k in _REVIEW_KEYS])

        # Wrap body column
        body_col = get_column_letter(_REVIEW_KEYS.index("body") + 1)
        for row_idx in range(2, ws.max_row + 1):
            ws[f"{body_col}{row_idx}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[1].height = 20
        _autowidth(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _autowidth(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
